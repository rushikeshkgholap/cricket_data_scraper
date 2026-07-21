"""
CricHeroes Full Automation Dashboard
====================================
Ek ya multiple CricHeroes scorecard links do -> ye automatically:
  1) Har match ka scorecard scrape karta hai (batting/bowling/extras/yet-to-bat/FOW)
  2) Har match se unique players (id/name/batting-hand) nikalta hai
  3) Har unique player ka profile page visit karke Batting_Style/Bowling_Style/Photo scrape karta hai
  4) Final Match_Master, Batting_Final, Bowling_Final, Players files banata hai
  5) Sab kuch Streamlit dashboard me dikhata hai + download button deta hai

CHALANE KA TARIKA (sirf apne local computer pe, Chrome installed hona chahiye):
    pip install streamlit pandas numpy requests beautifulsoup4 openpyxl pillow undetected-chromedriver selenium
    streamlit run app.py

NOTE: Streamlit Cloud / server pe Selenium+Chrome nahi chalega — ye sirf local machine pe chalega
jahan real Chrome browser installed hai. CHROME_MAIN_VERSION apne Chrome ke version se match karo
(chrome://settings/help me check karo).
"""

import os
import re
import time
import json
import shutil
import sys
import types
from collections import OrderedDict
from io import BytesIO

# --- compatibility shim: newer Python (3.12+) removed 'distutils',
# but undetected_chromedriver still imports from it. Same fix used in the lead-gen app. ---
if 'distutils' not in sys.modules:
    try:
        import distutils  # noqa: F401
    except ImportError:
        from packaging.version import Version as _Version

        class LooseVersion(_Version):
            def __init__(self, vstring):
                cleaned = str(vstring).split('-')[0]
                super().__init__(cleaned)

        distutils_module = types.ModuleType('distutils')
        version_module = types.ModuleType('distutils.version')
        version_module.LooseVersion = LooseVersion
        distutils_module.version = version_module
        sys.modules['distutils'] = distutils_module
        sys.modules['distutils.version'] = version_module

import numpy as np
import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

try:
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    SELENIUM_AVAILABLE = True
except Exception:
    # Streamlit Cloud jaise environments me Chrome browser nahi hota,
    # isliye scraping part disable ho jaayega, dashboard part chalta rahega
    SELENIUM_AVAILABLE = False


# =====================================================================
# CONFIG
# =====================================================================
CHROME_MAIN_VERSION = 148  # <-- apna Chrome version yahan daalo (chrome://settings/help)
MATCH_TYPE_KEYWORDS = ["Limited Overs", "Test", "T20", "T10", "ODI", "Box Cricket", "Pair Cricket"]

OUTPUT_DIR = "cricket_output"
IMAGE_FOLDER = os.path.join(OUTPUT_DIR, "temp_photos")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(IMAGE_FOLDER, exist_ok=True)

MATCH_MASTER_CSV = os.path.join(OUTPUT_DIR, "Match_Master.csv")
BATTING_CSV = os.path.join(OUTPUT_DIR, "Batting_Final.csv")
BOWLING_CSV = os.path.join(OUTPUT_DIR, "Bowling_Final.csv")
PLAYERS_CSV = os.path.join(OUTPUT_DIR, "Players.csv")
BATTING_XLSX = os.path.join(OUTPUT_DIR, "Batting_Final.xlsx")
BOWLING_XLSX = os.path.join(OUTPUT_DIR, "Bowling_Final.xlsx")

BAT_DROP = ["player_id", "is_impact_player_in", "is_impact_player_out", "minutes",
            "highlight_videos", "wicket_videos", "short_name", "type_code",
            "how_to_out_short_name", "batting_hand"]
BOWL_DROP = ["player_id", "is_impact_player_in", "is_impact_player_out", "minutes",
             "highlight_videos", "wicket_videos", "short_name", "type_code", "bowling_style"]


# =====================================================================
# BROWSER
# =====================================================================
try:
    from seleniumwire import webdriver as wire_webdriver
    SELENIUMWIRE_AVAILABLE = True
except Exception:
    SELENIUMWIRE_AVAILABLE = False


def _get_secret(key, default=""):
    """Streamlit secrets se proxy config uthao (cloud pe Settings > Secrets se aayega,
    local pe .streamlit/secrets.toml se). Agar kahin na mile to env var ya default use hoga."""
    try:
        return st.secrets["proxy"][key]
    except Exception:
        return os.environ.get(f"PROXY_{key.upper()}", default)


PROXY_HOST = _get_secret("host", "brd.superproxy.io")
PROXY_PORT = _get_secret("port", "33335")
PROXY_USER = _get_secret("user", "")
PROXY_PASS = _get_secret("pass", "")
ZENROWS_API_KEY = _get_secret("zenrows_api_key", "") or os.environ.get("ZENROWS_API_KEY", "")
USE_ZENROWS = bool(ZENROWS_API_KEY)


def fetch_html_via_zenrows(url):
    """ZenRows khud JS render + Cloudflare bypass (premium proxy) handle karta hai.
    Selenium/Chrome ki zaroorat hi nahi padti is path me."""
    params = {
        "url": url,
        "apikey": ZENROWS_API_KEY,
        "js_render": "true",
        "premium_proxy": "true",
    }
    resp = requests.get("https://api.zenrows.com/v1/", params=params, timeout=90)
    if resp.status_code != 200:
        raise RuntimeError(f"ZenRows error {resp.status_code}: {resp.text[:300]}")
    return resp.text


USE_PROXY = str(_get_secret("enabled", "true")).lower() == "true" and bool(PROXY_HOST) and bool(PROXY_PORT) and not USE_ZENROWS


def _check_proxy_ip(driver):
    """Diagnostic: confirm the proxy is actually being used by checking the outbound IP."""
    try:
        driver.get("https://geo.brdtest.com/welcome.txt")
        time.sleep(2)
        body = driver.find_element(By.TAG_NAME, "body").text[:300]
        return body
    except Exception as e:
        return f"(IP check failed: {e})"


def get_driver():
    if not SELENIUM_AVAILABLE:
        raise RuntimeError(
            "Selenium/Chrome yahan available nahi hai. Deploy karte waqt packages.txt "
            "(chromium + chromium-driver) commit karna mat bhoolo."
        )

    system_chromium = "/usr/bin/chromium" if os.path.exists("/usr/bin/chromium") else \
        ("/usr/bin/chromium-browser" if os.path.exists("/usr/bin/chromium-browser") else None)

    use_wire = USE_PROXY and SELENIUMWIRE_AVAILABLE

    if system_chromium:
        # ---- Streamlit Cloud / Linux server: headless system Chromium ----
        chromedriver_path = "/tmp/chromedriver"
        if not os.path.exists(chromedriver_path) and os.path.exists("/usr/bin/chromedriver"):
            shutil.copy("/usr/bin/chromedriver", chromedriver_path)
            os.chmod(chromedriver_path, 0o755)

        options = uc.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1920,1080")
        options.binary_location = system_chromium

        if use_wire:
            if PROXY_USER and PROXY_PASS:
                proxy_url = f"http://{PROXY_USER}:{PROXY_PASS}@{PROXY_HOST}:{PROXY_PORT}"
            else:
                proxy_url = f"http://{PROXY_HOST}:{PROXY_PORT}"  # free proxy: usually no auth needed
            sw_options = {
                "proxy": {"http": proxy_url, "https": proxy_url, "no_proxy": "localhost,127.0.0.1"}
            }
            driver = wire_webdriver.Chrome(
                options=options,
                seleniumwire_options=sw_options,
                service=uc.service.Service(executable_path=chromedriver_path) if hasattr(uc, "service") else None,
            )
        else:
            driver = uc.Chrome(options=options, driver_executable_path=chromedriver_path)
    else:
        # ---- Local machine: real Chrome, non-headless (matches your original notebook setup) ----
        options = uc.ChromeOptions()
        options.add_argument("--window-size=1920,1080")

        if use_wire:
            if PROXY_USER and PROXY_PASS:
                proxy_url = f"http://{PROXY_USER}:{PROXY_PASS}@{PROXY_HOST}:{PROXY_PORT}"
            else:
                proxy_url = f"http://{PROXY_HOST}:{PROXY_PORT}"
            sw_options = {
                "proxy": {"http": proxy_url, "https": proxy_url, "no_proxy": "localhost,127.0.0.1"}
            }
            driver = wire_webdriver.Chrome(options=options, seleniumwire_options=sw_options)
        else:
            driver = uc.Chrome(options=options, version_main=CHROME_MAIN_VERSION)

    return driver


# =====================================================================
# NEXT.JS RAW DATA + JSON EXTRACTION
# =====================================================================
def get_raw_blob(soup):
    combined = ""
    for s in soup.find_all("script"):
        txt = s.get_text()
        for m in re.finditer(r'self\.__next_f\.push\(\[1,"(.*?)"\]\)', txt, re.DOTALL):
            try:
                combined += m.group(1).encode("utf-8").decode("unicode_escape")
            except Exception:
                pass
    return combined if combined else None


def extract_json_array(raw, key):
    marker = f'"{key}":['
    start = raw.find(marker)
    if start == -1:
        return None
    i = start + len(marker) - 1
    depth = 0
    in_string = False
    escape = False
    j = i
    while j < len(raw):
        ch = raw[j]
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
        else:
            if ch == '"':
                in_string = True
            elif ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    return raw[i:j + 1]
        j += 1
    return None


def get_match_id(url):
    m = re.search(r'/scorecard/(\d+)/', url)
    return m.group(1) if m else None


# ---------- match details from rendered page text (fixed version) ----------
def get_match_text_fields(text, url):
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    header = ""
    header_idx = -1
    for idx, l in enumerate(lines):
        if l.startswith("Toss:"):
            break
        if any(kw in l for kw in MATCH_TYPE_KEYWORDS):
            header = l
            header_idx = idx
            break

    tournament = ""
    if header_idx > 0:
        tournament = re.sub(r'\s*\(.*?\)\s*$', '', lines[header_idx - 1]).strip()
    if not tournament:
        slug = url.split("/scorecard/")[-1].split("/")
        tournament = slug[1].replace("-", " ").title() if len(slug) > 1 else ""

    ground, match_type, overs, match_date = "", "", "", ""
    if header:
        for kw in MATCH_TYPE_KEYWORDS:
            if kw in header:
                ground = header.split(kw)[0].rstrip(", ").strip()
                rest = header.split(kw, 1)[1].lstrip(", ")
                parts = [p.strip() for p in rest.split(",") if p.strip()]
                match_type = kw
                overs = parts[0] if len(parts) > 0 else ""
                match_date = ",".join(parts[1:]).strip() if len(parts) > 1 else ""
                break

    toss_m = re.search(r"Toss:.*", text)
    toss = toss_m.group().strip() if toss_m else ""

    result_m = re.search(r".*won by.*|.*[Mm]atch [Dd]rawn.*|.*[Mm]atch tied.*", text)
    result = result_m.group().strip() if result_m else ""

    return {
        "tournament": tournament, "ground": ground, "match_type": match_type,
        "overs": overs, "date": match_date, "toss": toss, "result": result
    }


# =====================================================================
# DISMISSAL PARSER
# =====================================================================
def parse_dismissal(text):
    text = str(text).strip()
    lowered = text.lower()

    if not text or lowered in ["not out", "nan", "", "dnb"]:
        return ("Not Out" if lowered != "dnb" else "DNB"), "", ""
    if "retired hurt" in lowered:
        return "Retired Hurt", "", ""
    if "retired out" in lowered:
        return "Retired Out", "", ""

    text_clean = re.sub(r'[^\x00-\x7F]+', '', text)
    text_clean = text_clean.replace('â€ ', '').replace('†', '').strip()

    d_type, bowler, fielder = "Other", "", ""

    if "run out" in lowered:
        d_type = "Run Out"
        match = re.search(r'(?i)run out\s+(.*)', text_clean)
        if match:
            fielder = match.group(1).strip()
        return d_type, bowler, fielder
    elif lowered.startswith('c ') or ' c ' in lowered or 'c&b' in lowered or 'c & b' in lowered:
        d_type = "Caught"
        if 'c&b' in lowered or 'c & b' in lowered:
            match = re.search(r'(?i)(?:c\s*&\s*b|c&b)\s+(.*)', text_clean)
            if match:
                bowler = match.group(1).strip()
                fielder = bowler
        else:
            match = re.search(r'(?i)c\s+(.*?)\s+b\s+(.*)', text_clean)
            if match:
                fielder = match.group(1).strip()
                bowler = match.group(2).strip()
            else:
                match_f = re.search(r'(?i)c\s+(.*)', text_clean)
                if match_f:
                    fielder = match_f.group(1).strip()
        return d_type, bowler, fielder
    elif lowered.startswith('b '):
        d_type = "Bowled"
        bowler = text_clean[2:].strip()
        return d_type, bowler, fielder
    elif "lbw" in lowered:
        d_type = "LBW"
        match = re.search(r'(?i)lbw\s+b\s+(.*)', text_clean)
        if match:
            bowler = match.group(1).strip()
        return d_type, bowler, fielder
    elif lowered.startswith('st ') or ' st ' in lowered:
        d_type = "Stumped"
        match = re.search(r'(?i)st\s+(.*?)\s+b\s+(.*)', text_clean)
        if match:
            fielder = match.group(1).strip()
            bowler = match.group(2).strip()
        return d_type, bowler, fielder

    return d_type, bowler, fielder


def clean_name(name):
    if pd.isna(name):
        return name
    name = str(name)
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


# =====================================================================
# PLAYER ID / BATTING HAND HELPERS (raw JSON se)
# =====================================================================
def get_pid(p):
    if not isinstance(p, dict):
        return None
    for key in ["player_id", "playerId", "id", "user_id", "userId"]:
        if key in p and p.get(key) not in (None, ""):
            return p.get(key)
    return None


def get_bat_hand(p):
    if not isinstance(p, dict):
        return None
    for key in ["batting_hand", "battingHand", "batting_style"]:
        if key in p:
            return p.get(key)
    return None


def clean_batting(batters):
    df = pd.DataFrame(batters)
    if df.empty:
        return df
    df = df.rename(columns={"name": "Batter", "how_to_out": "Dismissal",
                             "runs": "Runs", "balls": "Balls"})
    return df.drop(columns=BAT_DROP, errors="ignore")


def clean_bowling(bowlers):
    df = pd.DataFrame(bowlers)
    if df.empty:
        return df
    df = df.rename(columns={"name": "Bowler", "economy_rate": "Economy"})
    return df.drop(columns=BOWL_DROP, errors="ignore")


# =====================================================================
# STAGE A: scrape one match scorecard
# =====================================================================
def scrape_match(url, driver, wait_seconds=15):
    match_id = get_match_id(url)

    if USE_ZENROWS:
        html = fetch_html_via_zenrows(url)
    else:
        driver.get(url)
        time.sleep(wait_seconds)
        html = driver.page_source

    soup = BeautifulSoup(html, "html.parser")

    raw = get_raw_blob(soup)
    if raw is None:
        page_title = (soup.title.string.strip() if soup.title and soup.title.string else "(no title)")
        html_len = len(html)
        lowered_html = html.lower()
        block_hint = ""
        for kw in ["captcha", "cloudflare", "access denied", "are you human",
                   "just a moment", "checking your browser", "unusual traffic"]:
            if kw in lowered_html:
                block_hint = f" — possible bot-block detected (found '{kw}' in page)"
                break
        return None, (
            f"Script data block nahi mila (page structure badal gaya ho sakta hai). "
            f"[debug: page_title='{page_title}', html_length={html_len}{block_hint}]"
        )

    scorecard_block = extract_json_array(raw, "scorecard")
    if scorecard_block is None:
        return None, "'scorecard' JSON array nahi mila."

    scorecard = json.loads(scorecard_block)
    if len(scorecard) < 2:
        return None, "Kam se kam 2 innings ka data nahi mila (match live/incomplete ho sakta hai)."

    if USE_ZENROWS:
        body_text = soup.get_text("\n")
    else:
        body_text = driver.find_element(By.TAG_NAME, "body").text
    md = get_match_text_fields(body_text, url)
    team_names = list(OrderedDict((inn.get("teamName", ""), None) for inn in scorecard).keys())

    return {
        "match_id": match_id, "url": url, "scorecard": scorecard,
        "match_details": md, "team_names": team_names,
    }, None


# =====================================================================
# STAGE B: build match/batting/bowling rows directly from scorecard JSON
# (xlsx round-trip ki zaroorat nahi — seedha memory se banate hain)
# =====================================================================
def build_match_tables(match_result):
    match_id = match_result["match_id"]
    md = match_result["match_details"]
    scorecard = match_result["scorecard"]
    team_names = match_result["team_names"]
    team1 = team_names[0] if len(team_names) > 0 else ""
    team2 = team_names[1] if len(team_names) > 1 else ""

    match_row = {
        "Match_ID": match_id, "Tournament": md["tournament"], "Ground": md["ground"],
        "Match Type": md["match_type"], "Overs": md["overs"], "Date": md["date"],
        "Toss": md["toss"], "Result": md["result"], "Team 1": team1, "Team 2": team2,
        "URL": match_result["url"],
    }

    all_batting_frames = []
    all_bowling_frames = []
    player_registry = {}  # player_id -> {Player_Name, Batting_Hand}

    for idx, inn in enumerate(scorecard):
        bat_team = inn.get("teamName", "")
        bowl_team = team2 if bat_team == team1 else team1

        batters_raw = inn.get("batting", [])
        bowlers_raw = inn.get("bowling", [])
        ytb_raw = inn.get("to_be_bat", [])

        # ---- Player_ID map (BEFORE columns get dropped by clean_batting) ----
        pid_by_name = {}
        for p in batters_raw + bowlers_raw + [x for x in ytb_raw if isinstance(x, dict)]:
            nm = p.get("name", "")
            pid = get_pid(p)
            if nm and pid is not None:
                pid_by_name[nm] = pid
            if pid is not None:
                player_registry.setdefault(pid, {
                    "Player_Name": nm,
                    "Batting_Hand": get_bat_hand(p)
                })

        # ---- Batting ----
        bat_df = clean_batting(batters_raw)
        if not bat_df.empty:
            bat_df["Player_ID"] = bat_df["Batter"].map(pid_by_name)
            parsed = bat_df["Dismissal"].apply(parse_dismissal)
            bat_df["Dismissal_Type"] = [x[0] for x in parsed]
            bat_df["Bowler"] = [x[1] for x in parsed]
            bat_df["Fielder"] = [x[2] for x in parsed]
            bat_df["Match_ID"] = match_id
            bat_df["Match"] = md["tournament"]
            bat_df["Inning_No"] = idx + 1
            bat_df["Team"] = bat_team

        # ---- Yet To Bat -> DNB rows, same column shape ----
        ytb_rows = []
        base_cols = list(bat_df.columns) if not bat_df.empty else \
            ["Batter", "Runs", "Balls", "4s", "6s", "SR", "Dismissal",
             "Player_ID", "Dismissal_Type", "Bowler", "Fielder"]
        for p in ytb_raw:
            nm = p.get("name", "") if isinstance(p, dict) else str(p)
            pid = get_pid(p) if isinstance(p, dict) else None
            row = {c: 0 for c in base_cols}
            row["Batter"] = nm
            row["Dismissal"] = "DNB"
            row["Dismissal_Type"] = "DNB"
            row["Bowler"] = ""
            row["Fielder"] = ""
            row["Player_ID"] = pid
            row["Match_ID"] = match_id
            row["Match"] = md["tournament"]
            row["Inning_No"] = idx + 1
            row["Team"] = bat_team
            if "SR" in row:
                row["SR"] = 0.0
            ytb_rows.append(row)

        if ytb_rows:
            bat_df = pd.concat([bat_df, pd.DataFrame(ytb_rows)], ignore_index=True) if not bat_df.empty \
                else pd.DataFrame(ytb_rows)

        if not bat_df.empty:
            all_batting_frames.append(bat_df)

        # ---- Bowling ----
        bowl_df = clean_bowling(bowlers_raw)
        if not bowl_df.empty:
            bowl_df["Player_ID"] = bowl_df["Bowler"].map(pid_by_name)
            bowl_df["Match_ID"] = match_id
            bowl_df["Match"] = md["tournament"]
            bowl_df["Inning_No"] = idx + 1
            bowl_df["Team"] = bowl_team
            all_bowling_frames.append(bowl_df)

    batting_df = pd.concat(all_batting_frames, ignore_index=True) if all_batting_frames else pd.DataFrame()
    bowling_df = pd.concat(all_bowling_frames, ignore_index=True) if all_bowling_frames else pd.DataFrame()

    if not batting_df.empty:
        batting_df["Batter"] = batting_df["Batter"].apply(clean_name)
    if not bowling_df.empty:
        bowling_df["Bowler"] = bowling_df["Bowler"].apply(clean_name)

    return match_row, batting_df, bowling_df, player_registry


# =====================================================================
# STAGE D: player profile scrape (Batting_Style / Bowling_Style / Photo)
# =====================================================================
def download_image_file(soup, player_id):
    try:
        img_tag = soup.find("img", attrs={"alt": lambda x: x and "profile" in x.lower()}) \
                  or soup.find("img", class_=lambda x: x and "profile" in x.lower())
        if img_tag and img_tag.get("src"):
            photo_url = img_tag.get("src")
            if photo_url.startswith("/"):
                photo_url = "https://cricheroes.com" + photo_url
            img_data = requests.get(photo_url, timeout=10).content
            temp_path = os.path.join(IMAGE_FOLDER, f"{player_id}.jpg")
            with open(temp_path, 'wb') as f:
                f.write(img_data)
            with PILImage.open(temp_path) as img:
                img = img.resize((80, 80), PILImage.Resampling.LANCZOS)
                img.save(temp_path)
            return temp_path, photo_url
    except Exception:
        pass
    return None, None


def scrape_player_profile(player_id, driver, wait_seconds=4):
    url = f"https://cricheroes.com/player-profile/{int(player_id)}/player"
    batting, bowling, local_img_path, online_photo_url = None, None, None, None
    for retry in range(2):
        try:
            if USE_ZENROWS:
                html = fetch_html_via_zenrows(url)
            else:
                driver.get(url)
                time.sleep(wait_seconds)
                html = driver.page_source
            soup = BeautifulSoup(html, "html.parser")
            text = soup.get_text("\n")
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            for idx, line in enumerate(lines):
                if line in ["RHB", "LHB"]:
                    batting = line
                    if idx + 1 < len(lines) and lines[idx + 1] not in ["Player", "matches", "stats"]:
                        bowling = lines[idx + 1]
                    break
            local_img_path, online_photo_url = download_image_file(soup, player_id)
            break
        except Exception:
            time.sleep(2)
    return batting, bowling, local_img_path, online_photo_url


# =====================================================================
# FINAL EXCEL BUILDERS (Batting_Final.xlsx / Bowling_Final.xlsx with photos)
# =====================================================================
def build_final_batting_excel(batting_df, players_df, out_path):
    players_subset = players_df[["Player_ID", "Player_Name", "Batting_Style", "Bowling_Style", "Profile_Photo_URL"]].copy()
    merged = batting_df.merge(players_subset, on="Player_ID", how="left")

    desired_order = [
        "Match_ID", "Team", "Batter", "Player_ID", "Profile_Photo",
        "Batting_Style", "Bowling_Style", "Runs", "Balls", "4s", "6s",
        "SR", "Dismissal", "Dismissal_Type", "Bowler", "Fielder", "Profile_Photo_URL"
    ]
    for c in merged.columns:
        if c not in desired_order and c != "Profile_Photo":
            desired_order.append(c)
    if "Profile_Photo" not in desired_order:
        desired_order.insert(4, "Profile_Photo")
    final_cols = [c for c in desired_order if c in merged.columns or c == "Profile_Photo"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batting Final Data"
    ws.append(final_cols)
    photo_col_idx = final_cols.index("Profile_Photo") + 1
    col_letter = get_column_letter(photo_col_idx)
    ws.column_dimensions[col_letter].width = 15

    for current_row, (_, row_data) in enumerate(merged.iterrows(), start=2):
        for col_idx, col_name in enumerate(final_cols, start=1):
            if col_name != "Profile_Photo":
                ws.cell(row=current_row, column=col_idx, value=row_data.get(col_name))
        ws.row_dimensions[current_row].height = 65
        p_id = row_data.get("Player_ID")
        if pd.notna(p_id):
            img_file = os.path.join(IMAGE_FOLDER, f"{int(p_id)}.jpg")
            if os.path.exists(img_file):
                try:
                    ws.add_image(OpenpyxlImage(img_file), f"{col_letter}{current_row}")
                except Exception:
                    pass
    wb.save(out_path)
    return merged


def build_final_bowling_excel(bowling_df, players_df, out_path):
    players_subset = players_df[["Player_ID", "Player_Name", "Batting_Style", "Bowling_Style", "Profile_Photo_URL"]].copy()
    merged = bowling_df.merge(players_subset, on="Player_ID", how="left")

    desired_order = [
        "Match_ID", "Team", "Bowler", "Player_ID", "Profile_Photo",
        "Batting_Style", "Bowling_Style", "overs", "balls", "maidens",
        "runs", "wickets", "0s", "4s", "6s", "wide", "noball", "Economy", "Profile_Photo_URL",
    ]
    for c in merged.columns:
        if c not in desired_order and c != "Profile_Photo":
            desired_order.append(c)
    if "Profile_Photo" not in desired_order:
        desired_order.insert(4, "Profile_Photo")
    final_cols = [c for c in desired_order if c in merged.columns or c == "Profile_Photo"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bowling Final Data"
    ws.append(final_cols)
    photo_col_idx = final_cols.index("Profile_Photo") + 1
    col_letter = get_column_letter(photo_col_idx)
    ws.column_dimensions[col_letter].width = 15

    for current_row, (_, row_data) in enumerate(merged.iterrows(), start=2):
        for col_idx, col_name in enumerate(final_cols, start=1):
            if col_name != "Profile_Photo":
                ws.cell(row=current_row, column=col_idx, value=row_data.get(col_name))
        ws.row_dimensions[current_row].height = 65
        p_id = row_data.get("Player_ID")
        if pd.notna(p_id):
            img_file = os.path.join(IMAGE_FOLDER, f"{int(p_id)}.jpg")
            if os.path.exists(img_file):
                try:
                    ws.add_image(OpenpyxlImage(img_file), f"{col_letter}{current_row}")
                except Exception:
                    pass
    wb.save(out_path)
    return merged


# =====================================================================
# FULL PIPELINE (runs inside Streamlit button callback)
# =====================================================================
def run_pipeline(links, scrape_photos, progress_cb):
    log_lines = []

    def log_and_progress(msg, frac):
        log_lines.append(msg)
        progress_cb(msg, frac)

    if USE_ZENROWS:
        log_and_progress("ZenRows API use ho rahi hai — Selenium/Chrome ki zaroorat nahi hai.", 0.0)
        driver = None
    else:
        driver = get_driver()
        if USE_PROXY:
            log_and_progress("Proxy check kar rahe hain...", 0.0)
            ip_info = _check_proxy_ip(driver)
            log_and_progress(f"Proxy IP check result: {ip_info}", 0.02)
        else:
            log_and_progress("Proxy OFF hai (USE_PROXY False ya credentials missing) — direct connection use ho rahi hai.", 0.0)

    match_rows = []
    batting_frames = []
    bowling_frames = []
    player_registry = {}  # player_id -> {Player_Name, Batting_Hand}
    errors = []

    total = len(links)
    try:
        # ---- Phase 1: scrape all match scorecards ----
        for i, link in enumerate(links, start=1):
            link = link.rstrip("/").replace("/summary", "/scorecard")
            log_and_progress(f"[{i}/{total}] Scraping match: {link}", (i - 1) / (total * 2))
            try:
                result, err = scrape_match(link, driver)
                if result is None:
                    errors.append(f"{link} -> {err}")
                    continue
                match_row, bat_df, bowl_df, p_reg = build_match_tables(result)
                match_rows.append(match_row)
                if not bat_df.empty:
                    batting_frames.append(bat_df)
                if not bowl_df.empty:
                    bowling_frames.append(bowl_df)
                for pid, info in p_reg.items():
                    player_registry.setdefault(pid, info)

                # incremental save (partial-run recovery)
                pd.DataFrame(match_rows).to_csv(MATCH_MASTER_CSV, index=False)
                if batting_frames:
                    pd.concat(batting_frames, ignore_index=True).to_csv(BATTING_CSV, index=False)
                if bowling_frames:
                    pd.concat(bowling_frames, ignore_index=True).to_csv(BOWLING_CSV, index=False)
            except Exception as e:
                errors.append(f"{link} -> {e}")

        batting_df = pd.concat(batting_frames, ignore_index=True) if batting_frames else pd.DataFrame()
        bowling_df = pd.concat(bowling_frames, ignore_index=True) if bowling_frames else pd.DataFrame()
        match_df = pd.DataFrame(match_rows)

        # ---- Phase 2: player profile scrape ----
        players_final = []
        pids = list(player_registry.keys())
        total_players = len(pids)
        for j, pid in enumerate(pids, start=1):
            info = player_registry[pid]
            log_and_progress(f"[{j}/{total_players}] Player profile: {info.get('Player_Name')}",
                        0.5 + (j - 1) / (max(total_players, 1) * 2))
            batting_style, bowling_style, img_path, photo_url = (None, None, None, None)
            if scrape_photos:
                try:
                    batting_style, bowling_style, img_path, photo_url = scrape_player_profile(pid, driver)
                except Exception as e:
                    errors.append(f"player {pid} -> {e}")
            players_final.append({
                "Player_ID": pid,
                "Player_Name": clean_name(info.get("Player_Name", "")),
                "Batting_Style": batting_style,
                "Bowling_Style": bowling_style,
                "Profile_Photo_URL": photo_url,
            })
            # incremental save
            pd.DataFrame(players_final).to_csv(PLAYERS_CSV, index=False, encoding="utf-8-sig")

        players_df = pd.DataFrame(players_final) if players_final else pd.DataFrame(
            columns=["Player_ID", "Player_Name", "Batting_Style", "Bowling_Style", "Profile_Photo_URL"])

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # ---- Phase 3: build final excel outputs ----
    log_and_progress("Building final Batting_Final.xlsx / Bowling_Final.xlsx ...", 0.95)
    batting_merged = pd.DataFrame()
    bowling_merged = pd.DataFrame()
    if not batting_df.empty and not players_df.empty:
        batting_merged = build_final_batting_excel(batting_df, players_df, BATTING_XLSX)
    if not bowling_df.empty and not players_df.empty:
        bowling_merged = build_final_bowling_excel(bowling_df, players_df, BOWLING_XLSX)

    log_and_progress("Done!", 1.0)

    return {
        "match_df": match_df,
        "batting_df": batting_merged if not batting_merged.empty else batting_df,
        "bowling_df": bowling_merged if not bowling_merged.empty else bowling_df,
        "players_df": players_df,
        "errors": errors,
        "log_lines": log_lines,
    }


# =====================================================================
# STREAMLIT UI
# =====================================================================
st.set_page_config(page_title="CricHeroes Full Automation", layout="wide")
st.title("🏏 CricHeroes Full Automation Dashboard")

CAN_SCRAPE = SELENIUM_AVAILABLE or USE_ZENROWS

if CAN_SCRAPE:
    st.caption("Ek ya multiple CricHeroes scorecard links do — scraping se lekar final Batting/Bowling/Match files tak, sab automatic.")
    if USE_ZENROWS:
        st.caption("🔑 ZenRows API active hai — Cloudflare bypass isi se ho raha hai.")
else:
    st.warning(
        "⚠️ Is environment me Chrome/Selenium available nahi hai aur ZenRows API key bhi set nahi hai, "
        "isliye live scraping yahan nahi chalegi. Ya to ZenRows API key secrets me daalo, ya scraping apne "
        "**local computer** pe `streamlit run app.py` se chalao, phir yahan neeche pehle-se-scraped files "
        "upload karo ya `cricket_output/` folder GitHub repo me commit kar do — wo automatically load ho jaayengi."
    )

# ---- Auto-load previously saved output files (e.g. committed to the GitHub repo) ----
if "results" not in st.session_state:
    if os.path.exists(MATCH_MASTER_CSV):
        auto_results = {
            "match_df": pd.read_csv(MATCH_MASTER_CSV) if os.path.exists(MATCH_MASTER_CSV) else pd.DataFrame(),
            "batting_df": pd.read_csv(BATTING_CSV) if os.path.exists(BATTING_CSV) else pd.DataFrame(),
            "bowling_df": pd.read_csv(BOWLING_CSV) if os.path.exists(BOWLING_CSV) else pd.DataFrame(),
            "players_df": pd.read_csv(PLAYERS_CSV) if os.path.exists(PLAYERS_CSV) else pd.DataFrame(),
            "errors": [],
        }
        st.session_state["results"] = auto_results

# ---- Scraping form: only usable where Selenium/Chrome OR ZenRows is actually available ----
with st.form("scrape_form"):
    links_text = st.text_area(
        "CricHeroes scorecard links (ek line me ek link):",
        height=150,
        placeholder="https://cricheroes.com/scorecard/26163958/30-yca-series-match/vijay-u-19-vs-cfn-u-19/scorecard",
        disabled=not CAN_SCRAPE,
    )
    scrape_photos = st.checkbox("Player profile photos + batting/bowling style bhi scrape karein (dheema hai)",
                                 value=True, disabled=not CAN_SCRAPE)
    submitted = st.form_submit_button("🚀 Run Pipeline", disabled=not CAN_SCRAPE)

if submitted and CAN_SCRAPE:
    links = [l.strip() for l in links_text.split("\n") if l.strip()]
    if not links:
        st.error("Kam se kam ek link daalo.")
    else:
        status_box = st.empty()
        progress_bar = st.progress(0)

        def progress_cb(msg, frac):
            status_box.info(msg)
            progress_bar.progress(min(max(frac, 0.0), 1.0))

        if USE_ZENROWS:
            spinner_msg = "Pipeline chal raha hai (ZenRows API ke zariye — koi browser nahi khulega, yahi normal hai)..."
        elif os.path.exists("/usr/bin/chromium"):
            spinner_msg = "Pipeline chal raha hai (headless mode — koi visible window nahi khulegi, yahi normal hai)..."
        else:
            spinner_msg = "Pipeline chal raha hai... browser window open hogi, use band mat karo."
        with st.spinner(spinner_msg):
            results = run_pipeline(links, scrape_photos, progress_cb)

        st.success("Pipeline complete!")

        with st.expander("📜 Pipeline Log (proxy check yahan dekho)", expanded=True):
            for line in results.get("log_lines", []):
                st.write("-", line)

        if results["errors"]:
            with st.expander(f"⚠️ {len(results['errors'])} error(s) aayi"):
                for e in results["errors"]:
                    st.write("-", e)

        st.session_state["results"] = results

# ---- Cloud-friendly manual upload path ----
if not CAN_SCRAPE:
    st.subheader("📤 Ya phir apni pehle-se-scraped files yahan upload karo")
    up_cols = st.columns(4)
    up_match = up_cols[0].file_uploader("Match_Master.csv", type=["csv"], key="up_match")
    up_bat = up_cols[1].file_uploader("Batting_Final (csv/xlsx)", type=["csv", "xlsx"], key="up_bat")
    up_bowl = up_cols[2].file_uploader("Bowling_Final (csv/xlsx)", type=["csv", "xlsx"], key="up_bowl")
    up_players = up_cols[3].file_uploader("Players.csv", type=["csv"], key="up_players")

    if st.button("📊 Load uploaded files into dashboard"):
        uploaded_results = st.session_state.get("results", {
            "match_df": pd.DataFrame(), "batting_df": pd.DataFrame(),
            "bowling_df": pd.DataFrame(), "players_df": pd.DataFrame(), "errors": []
        })
        if up_match is not None:
            uploaded_results["match_df"] = pd.read_csv(up_match)
        if up_bat is not None:
            uploaded_results["batting_df"] = pd.read_csv(up_bat) if up_bat.name.endswith(".csv") else pd.read_excel(up_bat)
        if up_bowl is not None:
            uploaded_results["bowling_df"] = pd.read_csv(up_bowl) if up_bowl.name.endswith(".csv") else pd.read_excel(up_bowl)
        if up_players is not None:
            uploaded_results["players_df"] = pd.read_csv(up_players)
        st.session_state["results"] = uploaded_results
        st.success("Files load ho gayi — neeche dashboard me dekho.")

# ---- Show results if available ----
if "results" in st.session_state:
    results = st.session_state["results"]

    tab1, tab2, tab3, tab4 = st.tabs(["📋 Match Master", "🏏 Batting Final", "🎯 Bowling Final", "👤 Players"])

    with tab1:
        st.dataframe(results["match_df"], use_container_width=True)
        if os.path.exists(MATCH_MASTER_CSV):
            st.download_button("⬇️ Download Match_Master.csv", open(MATCH_MASTER_CSV, "rb"),
                                file_name="Match_Master.csv")

    with tab2:
        st.dataframe(results["batting_df"], use_container_width=True)
        if os.path.exists(BATTING_XLSX):
            st.download_button("⬇️ Download Batting_Final.xlsx", open(BATTING_XLSX, "rb"),
                                file_name="Batting_Final.xlsx")
        elif os.path.exists(BATTING_CSV):
            st.download_button("⬇️ Download Batting_Final.csv", open(BATTING_CSV, "rb"),
                                file_name="Batting_Final.csv")

    with tab3:
        st.dataframe(results["bowling_df"], use_container_width=True)
        if os.path.exists(BOWLING_XLSX):
            st.download_button("⬇️ Download Bowling_Final.xlsx", open(BOWLING_XLSX, "rb"),
                                file_name="Bowling_Final.xlsx")
        elif os.path.exists(BOWLING_CSV):
            st.download_button("⬇️ Download Bowling_Final.csv", open(BOWLING_CSV, "rb"),
                                file_name="Bowling_Final.csv")

    with tab4:
        st.dataframe(results["players_df"], use_container_width=True)
        if os.path.exists(PLAYERS_CSV):
            st.download_button("⬇️ Download Players.csv", open(PLAYERS_CSV, "rb"),
                                file_name="Players.csv")
